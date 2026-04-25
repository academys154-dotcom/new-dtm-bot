import asyncio
import html
import json
import logging
import os
import re
import sqlite3
from contextlib import closing
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Iterable, Optional

from aiogram import Bot, Dispatcher, F, Router
from aiogram.client.default import DefaultBotProperties
from aiogram.enums import ContentType, ParseMode
from aiogram.exceptions import TelegramBadRequest, TelegramRetryAfter
from aiogram.filters import Command, CommandStart
from aiogram.fsm.context import FSMContext
from aiogram.fsm.state import State, StatesGroup
from aiogram.types import FSInputFile, KeyboardButton, Message, ReplyKeyboardMarkup, ReplyKeyboardRemove
from dotenv import load_dotenv
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

BASE_DIR = Path(__file__).resolve().parent.parent
load_dotenv(BASE_DIR / ".env")

logging.basicConfig(level=logging.INFO, format="%(asctime)s | %(levelname)s | %(name)s | %(message)s")
logger = logging.getLogger("dtm_school_exam_bot")

BTN_FULL_MOCK = "🎯 Full mock"
BTN_SUBJECTS = "📚 Fanlar"
BTN_CHECK = "✅ Test tekshirish"
BTN_PROFILE = "👤 Shaxsiy ma'lumotlar"
BTN_GUIDE = "🧭 Qo'llanma"
BTN_HELP = "🆘 Yordam"
BTN_CLEAR = "🧹 Clear all"
BTN_ADMIN = "🛠 Admin panel"

BTN_BACK = "⬅️ Orqaga"
BTN_HOME = "🏠 Asosiy menu"
BTN_ENTER_CODE = "🔑 Kod kiritish"
BTN_FREE_TEST = "🎁 Bir martalik tekin test"
BTN_TESTS = "📄 Testlar ro'yxati"
FULL_MOCK_BONUS_SUBJECT = "__FULL_MOCK_BONUS__"
FULL_MOCK_BONUS_GRADE = "GLOBAL"

BTN_EXAM_ADD = "➕ Test qo'shish"
BTN_EXAM_DELETE = "🗑 Testni o'chirish"
BTN_EXAM_LIST = "📋 Testlar ro'yxati"
BTN_CODE_CREATE = "🔐 Kod yaratish"
BTN_STATS = "📈 Statistika"
BTN_EXPORT_USERS = "📤 Userlar export"
BTN_EXPORT_RESULTS = "📥 Natijalar export"
BTN_BROADCAST = "📣 Eslatma"
BTN_ADMIN_TO_USER = "👥 User menu"

CATEGORY_EXACT = "📐 Aniq fanlar"
CATEGORY_NATURAL = "🧪 Tabiiy fanlar"
CATEGORY_SOCIAL = "🏛 Ijtimoiy fanlar"

SUBJ_MATH = "Matematika"
SUBJ_ENGLISH = "Ingliz tili"
SUBJ_NATIVE = "Ona tili"
SUBJ_HISTORY = "Tarix"
SUBJ_GEOGRAPHY = "Geografiya"
SUBJ_BIOLOGY = "Biologiya"
SUBJ_CHEMISTRY = "Kimyo"
SUBJ_PHYSICS = "Fizika"

SUBJECTS = [
    SUBJ_MATH,
    SUBJ_ENGLISH,
    SUBJ_NATIVE,
    SUBJ_HISTORY,
    SUBJ_GEOGRAPHY,
    SUBJ_BIOLOGY,
    SUBJ_CHEMISTRY,
    SUBJ_PHYSICS,
]

CATEGORY_SUBJECTS = {
    CATEGORY_EXACT: [SUBJ_MATH, SUBJ_PHYSICS, SUBJ_ENGLISH],
    CATEGORY_NATURAL: [SUBJ_BIOLOGY, SUBJ_CHEMISTRY, SUBJ_GEOGRAPHY],
    CATEGORY_SOCIAL: [SUBJ_NATIVE, SUBJ_HISTORY],
}
SUBJECT_TO_CATEGORY = {s: cat for cat, items in CATEGORY_SUBJECTS.items() for s in items}
GRADE_BUTTONS = {"9-sinf": "9", "11-sinf": "11", "Hozircha bilmayman": "UNKNOWN"}
RESULT_DETAIL_LIMIT = 35
DEFAULT_DURATION = 180
DEFAULT_HELP_TEXT = "Yordam uchun admin bilan bog'laning."

INTERRUPTIBLE_USER_TEXTS = {
    BTN_FULL_MOCK, BTN_SUBJECTS, BTN_CHECK, BTN_PROFILE, BTN_GUIDE, BTN_HELP, BTN_CLEAR, BTN_ADMIN,
    BTN_BACK, BTN_HOME, BTN_ENTER_CODE, BTN_FREE_TEST, BTN_TESTS,
    BTN_EXAM_ADD, BTN_EXAM_DELETE, BTN_EXAM_LIST, BTN_CODE_CREATE, BTN_STATS, BTN_EXPORT_USERS, BTN_EXPORT_RESULTS, BTN_BROADCAST, BTN_ADMIN_TO_USER,
    *CATEGORY_SUBJECTS.keys(), *SUBJECTS,
}

ADMIN_MENU_OPTIONS = [
    BTN_EXAM_ADD,
    BTN_EXAM_DELETE,
    BTN_EXAM_LIST,
    BTN_CODE_CREATE,
    BTN_STATS,
    BTN_BROADCAST,
    BTN_EXPORT_USERS,
    BTN_EXPORT_RESULTS,
    BTN_ADMIN_TO_USER,
    BTN_CLEAR,
]

USER_MENU_OPTIONS = [
    BTN_FULL_MOCK,
    BTN_SUBJECTS,
    BTN_CHECK,
    BTN_PROFILE,
    BTN_GUIDE,
    BTN_HELP,
    BTN_CLEAR,
]

GRADE_OPTIONS = [("9-sinf", "9"), ("11-sinf", "11"), ("Barcha sinflar", "ALL")]
EXPORT_FORMAT_OPTIONS = [("XLSX", "1"), ("PDF", "2")]
CODE_MODE_OPTIONS = [("Bitta kod", "1"), ("Ko'p kod", "2")]
YES_NO_OPTIONS = [("Ha", "yes"), ("Yo'q", "no")]


@dataclass(frozen=True)
class Settings:
    bot_token: str
    admin_ids: set[int]
    db_path: str
    bot_brand_name: str
    help_admin_url: str
    help_admin_label: str


class RegisterStates(StatesGroup):
    waiting_name = State()
    waiting_contact = State()
    waiting_grade = State()


class UserStates(StatesGroup):
    waiting_redeem_code = State()
    waiting_check_code = State()
    waiting_answers = State()


class AdminStates(StatesGroup):
    waiting_exam_title = State()
    waiting_exam_grade = State()
    waiting_exam_category = State()
    waiting_exam_subject = State()
    waiting_exam_duration = State()
    waiting_exam_description = State()
    waiting_exam_free_flag = State()
    waiting_exam_pdf = State()
    waiting_exam_answer_key = State()
    waiting_exam_instructions = State()
    waiting_delete_exam_id = State()
    waiting_code_exam_id = State()
    waiting_code_mode = State()
    waiting_single_code = State()
    waiting_bulk_codes = State()
    waiting_export_users_format = State()
    waiting_export_results_format = State()
    waiting_broadcast_text = State()


def utc_now() -> str:
    return datetime.now(timezone.utc).isoformat(timespec="seconds")


def compact_spaces(text: str) -> str:
    return re.sub(r"\s+", " ", text or "").strip()


def numbered_list(options: list[str]) -> str:
    return "\n".join(f"{index}) {option}" for index, option in enumerate(options, start=1))


def numbered_prompt(title: str, options: list[str]) -> str:
    return f"{title}\n\n{numbered_list(options)}\n\nRaqamini yuborsangiz ham bo'ladi. Masalan: 2"


def pick_numbered_value(text: str, options: list[str]) -> str:
    raw = compact_spaces(text)
    if raw.isdigit():
        index = int(raw) - 1
        if 0 <= index < len(options):
            return options[index]
    return raw if raw in options else ""


def pick_labeled_value(text: str, options: list[tuple[str, str]]) -> str:
    raw = compact_spaces(text)
    if raw.isdigit():
        index = int(raw) - 1
        if 0 <= index < len(options):
            return options[index][1]
    lowered = raw.lower()
    for label, value in options:
        if raw == label or lowered == label.lower() or lowered == value.lower():
            return value
    return ""


def normalize_code(text: str) -> str:
    return re.sub(r"[^A-Za-z0-9_-]", "", (text or "").strip().upper())


def normalize_grade_value(text: str) -> str:
    raw = compact_spaces(text).lower()
    if raw in {"9", "9-sinf", "9 sinf", "sinf 9"}:
        return "9"
    if raw in {"11", "11-sinf", "11 sinf", "sinf 11"}:
        return "11"
    if raw in {"all", "hammasi", "barchasi", "umumiy", "0"}:
        return "ALL"
    if raw in {"bilmayman", "hozircha bilmayman", "unknown", "-"}:
        return "UNKNOWN"
    return ""


def grade_label(value: str) -> str:
    return {"9": "9-sinf", "11": "11-sinf", "ALL": "Barcha sinflar", "UNKNOWN": "Ko'rsatilmagan"}.get(value or "", value or "Ko'rsatilmagan")


def is_valid_phone(phone: str) -> bool:
    digits = re.sub(r"\D", "", phone or "")
    return 9 <= len(digits) <= 15


def is_valid_full_name(full_name: str) -> bool:
    full_name = compact_spaces(full_name)
    if len(full_name) < 3 or len(full_name) > 100:
        return False
    if re.search(r"\d", full_name):
        return False
    return bool(re.fullmatch(r"[A-Za-zÀ-ÿĀ-žА-Яа-яЁёʻ'’‘\-\s.]+", full_name))


def is_pdf_document(message: Message) -> bool:
    document = message.document
    if not document:
        return False
    mime_type = (document.mime_type or "").lower()
    file_name = (document.file_name or "").lower()
    return mime_type == "application/pdf" or file_name.endswith(".pdf")


def resolve_db_path(raw_path: str) -> str:
    candidate = Path(raw_path or (BASE_DIR / "data/data.sqlite3")).expanduser()
    if not candidate.is_absolute():
        candidate = BASE_DIR / candidate
    return str(candidate.resolve())


def ensure_parent_dir(file_path: str) -> None:
    Path(file_path).expanduser().resolve().parent.mkdir(parents=True, exist_ok=True)


def clean_token(token: str) -> str:
    token = (token or "").strip().upper()
    token = token.replace("ʻ", "'").replace("’", "'").replace("‘", "'")
    token = re.sub(r"[^A-Z0-9+'-]", "", token)
    return token or "-"


def parse_answer_tokens(text: str) -> list[str]:
    raw = (text or "").upper().replace("\r", "\n")
    raw = raw.replace("ʻ", "'").replace("’", "'").replace("‘", "'")
    raw = raw.strip()
    if not raw:
        return []

    compact_letters = re.sub(r"[^A-D]", "", raw)
    remaining = re.sub(r"[A-D\s,;|/\\\-.0-9()_:=]", "", raw)
    if compact_letters and not remaining:
        return list(compact_letters)

    text_numbered = re.sub(r"(\d+)\s*[-.)=:]\s*", r"\1:", raw)
    numbered = re.findall(r"\b\d+\s*:\s*([A-Z0-9+'-]+)", text_numbered)
    if numbered:
        return [clean_token(x) for x in numbered]

    prepared = re.sub(r"[\n,;|/\\]+", " ", raw)
    prepared = re.sub(r"\s+", " ", prepared).strip()
    if not prepared:
        return []
    tokens = [clean_token(item) for item in prepared.split(" ") if item.strip()]
    return [t for t in tokens if t]


def serialize_tokens(tokens: list[str]) -> str:
    return json.dumps(tokens, ensure_ascii=False)


def deserialize_tokens(payload: str) -> list[str]:
    payload = (payload or "").strip()
    if not payload:
        return []
    try:
        value = json.loads(payload)
        if isinstance(value, list):
            return [clean_token(str(x)) for x in value]
    except json.JSONDecodeError:
        pass
    return parse_answer_tokens(payload)


def estimate_five_grade(percent: float) -> int:
    if percent >= 86:
        return 5
    if percent >= 71:
        return 4
    if percent >= 56:
        return 3
    return 2


def performance_band(percent: float) -> str:
    if percent >= 90:
        return "A'lo"
    if percent >= 75:
        return "Yaxshi"
    if percent >= 56:
        return "Qoniqarli"
    return "Yana ishlash kerak"


def evaluate_answers(answer_key_tokens: list[str], user_answers_tokens: list[str]) -> dict:
    key = [clean_token(x) for x in answer_key_tokens]
    given = [clean_token(x) for x in user_answers_tokens]
    total = len(key)
    correct = 0
    blank = 0
    detail_lines: list[str] = []
    for index, correct_answer in enumerate(key, start=1):
        user_answer = given[index - 1] if index - 1 < len(given) else "-"
        is_blank = user_answer == "-"
        is_ok = user_answer == correct_answer
        if is_ok:
            correct += 1
        elif is_blank:
            blank += 1
        if index <= RESULT_DETAIL_LIMIT:
            detail_lines.append(f"{index}. Siz: {user_answer} | To'g'ri: {correct_answer} | {'✅' if is_ok else '❌'}")
    wrong = max(total - correct - blank, 0)
    percent = round((correct / total) * 100, 1) if total else 0.0
    five_grade = estimate_five_grade(percent)
    lines = [
        "📄 Natija",
        f"To'g'ri: {correct}/{total}",
        f"Xato: {wrong}",
        f"Bo'sh: {blank}",
        f"Foiz: {percent}%",
        f"5 ballik taxminiy baho: {five_grade}",
        f"Holat: {performance_band(percent)}",
        "",
        "🔍 Tahlil",
        *detail_lines,
    ]
    if total > RESULT_DETAIL_LIMIT:
        lines.append(f"... yana {total - RESULT_DETAIL_LIMIT} ta savol bor.")
    return {
        "correct_count": correct,
        "wrong_count": wrong,
        "blank_count": blank,
        "total_questions": total,
        "percentage": percent,
        "five_grade": five_grade,
        "result_text": "\n".join(lines),
    }


def escape_html_text(text: str) -> str:
    return html.escape(text or "")


def find_pdf_font() -> str:
    common_paths = [
        "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf",
        "/usr/share/fonts/truetype/liberation2/LiberationSans-Regular.ttf",
        "C:/Windows/Fonts/arial.ttf",
        "/System/Library/Fonts/Supplemental/Arial.ttf",
    ]
    for path in common_paths:
        if os.path.exists(path):
            if "AppFont" not in pdfmetrics.getRegisteredFontNames():
                pdfmetrics.registerFont(TTFont("AppFont", path))
            return "AppFont"
    return "Helvetica"


def safe_pdf_text(value: object) -> str:
    text = str(value or "")
    return text.replace("ʻ", "'").replace("’", "'").replace("‘", "'")


def build_xlsx_file(filename: str, headers: list[str], rows: list[list[object]], title: str) -> Path:
    export_dir = BASE_DIR / "exports"
    export_dir.mkdir(parents=True, exist_ok=True)
    file_path = export_dir / filename
    wb = Workbook()
    ws = wb.active
    ws.title = "Export"
    ws.freeze_panes = "A3"
    ws.sheet_view.showGridLines = False
    ws["A1"] = title
    ws["A1"].font = Font(bold=True, size=14)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    thin_border = Border(bottom=Side(style="thin", color="D9D9D9"))
    for col_index, header in enumerate(headers, start=1):
        cell = ws.cell(row=2, column=col_index, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border
    for row_index, row_data in enumerate(rows, start=3):
        for col_index, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_index, column=col_index, value=value)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = thin_border
    for idx, header in enumerate(headers, start=1):
        max_len = len(str(header))
        for row in rows:
            if idx - 1 < len(row):
                max_len = max(max_len, len(str(row[idx - 1] or "")))
        ws.column_dimensions[get_column_letter(idx)].width = min(max(max_len + 2, 14), 35)
    wb.save(file_path)
    return file_path


def build_pdf_file(filename: str, title: str, headers: list[str], rows: list[list[object]]) -> Path:
    export_dir = BASE_DIR / "exports"
    export_dir.mkdir(parents=True, exist_ok=True)
    file_path = export_dir / filename
    font_name = find_pdf_font()
    doc = SimpleDocTemplate(str(file_path), pagesize=landscape(A4), leftMargin=10 * mm, rightMargin=10 * mm, topMargin=10 * mm, bottomMargin=10 * mm)
    styles = getSampleStyleSheet()
    normal_style = ParagraphStyle("AppNormal", parent=styles["Normal"], fontName=font_name, fontSize=8, leading=10)
    title_style = ParagraphStyle("AppTitle", parent=styles["Title"], fontName=font_name, fontSize=16, leading=18)
    flow = [Paragraph(safe_pdf_text(title), title_style), Spacer(1, 6)]
    table_data = [[Paragraph(safe_pdf_text(header), normal_style) for header in headers]]
    for row in rows:
        table_data.append([Paragraph(safe_pdf_text(value), normal_style) for value in row])
    table = Table(table_data, repeatRows=1)
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F4E78")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, -1), font_name),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("GRID", (0, 0), (-1, -1), 0.3, colors.HexColor("#D9D9D9")),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.whitesmoke, colors.HexColor("#F7FBFF")]),
        ("LEFTPADDING", (0, 0), (-1, -1), 4),
        ("RIGHTPADDING", (0, 0), (-1, -1), 4),
    ]))
    flow.append(table)
    doc.build(flow)
    return file_path


class Database:
    def __init__(self, db_path: str) -> None:
        self.db_path = db_path
        self._init_db()

    def connect(self) -> sqlite3.Connection:
        ensure_parent_dir(self.db_path)
        conn = sqlite3.connect(self.db_path, timeout=30)
        conn.row_factory = sqlite3.Row
        return conn

    def _column_exists(self, conn: sqlite3.Connection, table_name: str, column_name: str) -> bool:
        rows = conn.execute(f"PRAGMA table_info({table_name})").fetchall()
        return any(row[1] == column_name for row in rows)

    def _ensure_column(self, conn: sqlite3.Connection, table_name: str, column_name: str, column_sql: str) -> None:
        if not self._column_exists(conn, table_name, column_name):
            conn.execute(f"ALTER TABLE {table_name} ADD COLUMN {column_name} {column_sql}")

    def _init_db(self) -> None:
        with closing(self.connect()) as conn:
            cur = conn.cursor()
            cur.executescript(
                """
                PRAGMA journal_mode=WAL;

                CREATE TABLE IF NOT EXISTS users (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    tg_user_id INTEGER NOT NULL UNIQUE,
                    username TEXT,
                    full_name TEXT,
                    phone TEXT,
                    target_grade TEXT DEFAULT 'UNKNOWN',
                    registered_at TEXT NOT NULL
                );

                CREATE TABLE IF NOT EXISTS exams (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    title TEXT NOT NULL,
                    grade_level TEXT DEFAULT 'ALL',
                    category TEXT DEFAULT '',
                    subject TEXT DEFAULT '',
                    duration_minutes INTEGER DEFAULT 180,
                    description TEXT DEFAULT '',
                    pdf_file_id TEXT NOT NULL,
                    answer_key TEXT NOT NULL,
                    answer_instructions TEXT DEFAULT '',
                    is_free_demo INTEGER NOT NULL DEFAULT 0,
                    is_active INTEGER NOT NULL DEFAULT 1,
                    created_at TEXT NOT NULL
                );

                CREATE TABLE IF NOT EXISTS redeem_codes (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    code TEXT NOT NULL UNIQUE,
                    exam_id INTEGER NOT NULL,
                    created_by_admin INTEGER NOT NULL,
                    created_at TEXT NOT NULL,
                    is_used INTEGER NOT NULL DEFAULT 0,
                    used_by_tg_user_id INTEGER,
                    used_at TEXT,
                    FOREIGN KEY (exam_id) REFERENCES exams(id)
                );

                CREATE TABLE IF NOT EXISTS attempts (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    tg_user_id INTEGER NOT NULL,
                    exam_id INTEGER NOT NULL,
                    code TEXT NOT NULL,
                    user_answers TEXT NOT NULL,
                    score INTEGER NOT NULL,
                    total_questions INTEGER NOT NULL,
                    percentage REAL NOT NULL DEFAULT 0,
                    correct_count INTEGER NOT NULL DEFAULT 0,
                    wrong_count INTEGER NOT NULL DEFAULT 0,
                    blank_count INTEGER NOT NULL DEFAULT 0,
                    five_grade INTEGER NOT NULL DEFAULT 0,
                    result_text TEXT NOT NULL,
                    created_at TEXT NOT NULL
                );

                CREATE TABLE IF NOT EXISTS message_history (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    tg_user_id INTEGER NOT NULL,
                    chat_id INTEGER NOT NULL,
                    message_id INTEGER NOT NULL,
                    sender_type TEXT NOT NULL,
                    created_at TEXT NOT NULL
                );

                CREATE TABLE IF NOT EXISTS free_demo_usage (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    tg_user_id INTEGER NOT NULL,
                    exam_id INTEGER NOT NULL,
                    subject TEXT NOT NULL,
                    grade_level TEXT NOT NULL,
                    used_at TEXT NOT NULL,
                    UNIQUE(tg_user_id, subject, grade_level)
                );
                """
            )
            self._ensure_column(conn, "exams", "category", "TEXT DEFAULT ''")
            self._ensure_column(conn, "exams", "is_free_demo", "INTEGER NOT NULL DEFAULT 0")
            cur.executescript(
                """
                CREATE INDEX IF NOT EXISTS idx_users_tg_user_id ON users(tg_user_id);
                CREATE INDEX IF NOT EXISTS idx_redeem_codes_code ON redeem_codes(code);
                CREATE INDEX IF NOT EXISTS idx_redeem_codes_used_by ON redeem_codes(used_by_tg_user_id);
                CREATE INDEX IF NOT EXISTS idx_attempts_tg_user_id ON attempts(tg_user_id);
                CREATE INDEX IF NOT EXISTS idx_attempts_exam_id ON attempts(exam_id);
                CREATE INDEX IF NOT EXISTS idx_message_history_chat_user ON message_history(chat_id, tg_user_id);
                CREATE INDEX IF NOT EXISTS idx_free_demo_usage_user_subject ON free_demo_usage(tg_user_id, grade_level, subject);
                """
            )
            conn.commit()

    def upsert_user(self, tg_user_id: int, username: Optional[str] = None, full_name: Optional[str] = None, phone: Optional[str] = None, target_grade: Optional[str] = None) -> None:
        with closing(self.connect()) as conn:
            row = conn.execute("SELECT id FROM users WHERE tg_user_id = ?", (tg_user_id,)).fetchone()
            if row:
                fields, values = [], []
                if username is not None:
                    fields.append("username = ?")
                    values.append(username)
                if full_name is not None:
                    fields.append("full_name = ?")
                    values.append(full_name)
                if phone is not None:
                    fields.append("phone = ?")
                    values.append(phone)
                if target_grade is not None:
                    fields.append("target_grade = ?")
                    values.append(target_grade)
                if fields:
                    values.append(tg_user_id)
                    conn.execute(f"UPDATE users SET {', '.join(fields)} WHERE tg_user_id = ?", tuple(values))
            else:
                conn.execute(
                    "INSERT INTO users (tg_user_id, username, full_name, phone, target_grade, registered_at) VALUES (?, ?, ?, ?, ?, ?)",
                    (tg_user_id, username, full_name, phone, target_grade or "UNKNOWN", utc_now()),
                )
            conn.commit()

    def get_user(self, tg_user_id: int) -> Optional[sqlite3.Row]:
        with closing(self.connect()) as conn:
            return conn.execute("SELECT * FROM users WHERE tg_user_id = ?", (tg_user_id,)).fetchone()

    def list_users(self) -> list[sqlite3.Row]:
        with closing(self.connect()) as conn:
            return conn.execute(
                "SELECT u.*, (SELECT COUNT(*) FROM attempts a WHERE a.tg_user_id=u.tg_user_id) as attempts_count FROM users u ORDER BY registered_at DESC"
            ).fetchall()

    def create_exam(self, title: str, grade_level: str, category: str, subject: str, duration_minutes: int, description: str, pdf_file_id: str, answer_key_tokens: list[str], answer_instructions: str, is_free_demo: bool) -> int:
        with closing(self.connect()) as conn:
            cur = conn.execute(
                """
                INSERT INTO exams (title, grade_level, category, subject, duration_minutes, description, pdf_file_id, answer_key, answer_instructions, is_free_demo, created_at)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (title.strip(), grade_level, category.strip(), subject.strip(), duration_minutes, description.strip(), pdf_file_id.strip(), serialize_tokens(answer_key_tokens), answer_instructions.strip(), 1 if is_free_demo else 0, utc_now()),
            )
            conn.commit()
            return int(cur.lastrowid)

    def list_exams(self, subject: Optional[str] = None, grade_level: Optional[str] = None, category: Optional[str] = None, include_free: bool = True) -> list[sqlite3.Row]:
        query = "SELECT * FROM exams WHERE is_active = 1"
        params: list[object] = []
        if subject:
            query += " AND subject = ?"
            params.append(subject)
        if category:
            query += " AND category = ?"
            params.append(category)
        if grade_level and grade_level not in {"UNKNOWN", "ALL"}:
            query += " AND (grade_level = ? OR grade_level = 'ALL')"
            params.append(grade_level)
        if not include_free:
            query += " AND is_free_demo = 0"
        query += " ORDER BY id DESC"
        with closing(self.connect()) as conn:
            return conn.execute(query, tuple(params)).fetchall()

    def get_exam(self, exam_id: int) -> Optional[sqlite3.Row]:
        with closing(self.connect()) as conn:
            return conn.execute("SELECT * FROM exams WHERE id = ? AND is_active = 1", (exam_id,)).fetchone()

    def deactivate_exam(self, exam_id: int) -> bool:
        with closing(self.connect()) as conn:
            cur = conn.execute("UPDATE exams SET is_active = 0 WHERE id = ?", (exam_id,))
            conn.commit()
            return cur.rowcount > 0

    def get_latest_free_demo_exam(self, grade_level: str, category: str, subject: str) -> Optional[sqlite3.Row]:
        with closing(self.connect()) as conn:
            return conn.execute(
                """
                SELECT * FROM exams
                WHERE is_active = 1
                  AND is_free_demo = 1
                  AND category = ?
                  AND subject = ?
                  AND (grade_level = ? OR grade_level = 'ALL')
                ORDER BY id DESC
                LIMIT 1
                """,
                (category, subject, grade_level),
            ).fetchone()

    def has_used_free_demo(self, tg_user_id: int, grade_level: str, subject: str) -> bool:
        with closing(self.connect()) as conn:
            row = conn.execute(
                "SELECT 1 FROM free_demo_usage WHERE tg_user_id = ? AND grade_level = ? AND subject = ? LIMIT 1",
                (tg_user_id, grade_level, subject),
            ).fetchone()
            return bool(row)

    def mark_free_demo_used(self, tg_user_id: int, exam_id: int, grade_level: str, subject: str) -> None:
        with closing(self.connect()) as conn:
            conn.execute(
                "INSERT OR IGNORE INTO free_demo_usage (tg_user_id, exam_id, subject, grade_level, used_at) VALUES (?, ?, ?, ?, ?)",
                (tg_user_id, exam_id, subject, grade_level, utc_now()),
            )
            conn.commit()

    def has_used_full_mock_bonus(self, tg_user_id: int) -> bool:
        return self.has_used_free_demo(tg_user_id, FULL_MOCK_BONUS_GRADE, FULL_MOCK_BONUS_SUBJECT)

    def mark_full_mock_bonus_used(self, tg_user_id: int, exam_id: int) -> None:
        self.mark_free_demo_used(tg_user_id, exam_id, FULL_MOCK_BONUS_GRADE, FULL_MOCK_BONUS_SUBJECT)

    def get_latest_full_mock_bonus_exam(self, grade_level: str = "UNKNOWN") -> Optional[sqlite3.Row]:
        with closing(self.connect()) as conn:
            if grade_level in {"9", "11"}:
                return conn.execute(
                    """
                    SELECT * FROM exams
                    WHERE is_active = 1
                      AND is_free_demo = 1
                      AND (grade_level = ? OR grade_level = 'ALL')
                    ORDER BY id DESC
                    LIMIT 1
                    """,
                    (grade_level,),
                ).fetchone()
            return conn.execute(
                """
                SELECT * FROM exams
                WHERE is_active = 1
                  AND is_free_demo = 1
                ORDER BY id DESC
                LIMIT 1
                """
            ).fetchone()

    def list_user_codes(self, tg_user_id: int) -> list[sqlite3.Row]:
        with closing(self.connect()) as conn:
            return conn.execute(
                """
                SELECT rc.code, rc.used_at, e.title, e.subject, e.category, e.grade_level
                FROM redeem_codes rc
                JOIN exams e ON e.id = rc.exam_id
                WHERE rc.used_by_tg_user_id = ?
                ORDER BY rc.used_at DESC, rc.id DESC
                """,
                (tg_user_id,),
            ).fetchall()

    def create_code(self, code: str, exam_id: int, admin_id: int) -> bool:
        try:
            with closing(self.connect()) as conn:
                conn.execute(
                    "INSERT INTO redeem_codes (code, exam_id, created_by_admin, created_at) VALUES (?, ?, ?, ?)",
                    (normalize_code(code), exam_id, admin_id, utc_now()),
                )
                conn.commit()
                return True
        except sqlite3.IntegrityError:
            return False

    def bulk_create_codes(self, codes: Iterable[str], exam_id: int, admin_id: int) -> tuple[int, int]:
        created = skipped = 0
        with closing(self.connect()) as conn:
            for raw_code in codes:
                code = normalize_code(raw_code)
                if not code:
                    continue
                try:
                    conn.execute(
                        "INSERT INTO redeem_codes (code, exam_id, created_by_admin, created_at) VALUES (?, ?, ?, ?)",
                        (code, exam_id, admin_id, utc_now()),
                    )
                    created += 1
                except sqlite3.IntegrityError:
                    skipped += 1
            conn.commit()
        return created, skipped

    def get_code_with_exam(self, code: str) -> Optional[sqlite3.Row]:
        with closing(self.connect()) as conn:
            return conn.execute(
                """
                SELECT rc.*, e.title, e.grade_level, e.subject, e.category, e.duration_minutes, e.description,
                       e.pdf_file_id, e.answer_key, e.answer_instructions, e.is_free_demo
                FROM redeem_codes rc
                JOIN exams e ON e.id = rc.exam_id
                WHERE rc.code = ? AND e.is_active = 1
                """,
                (normalize_code(code),),
            ).fetchone()

    def redeem_code(self, code: str, tg_user_id: int) -> tuple[bool, str]:
        code = normalize_code(code)
        with closing(self.connect()) as conn:
            row = conn.execute("SELECT * FROM redeem_codes WHERE code = ?", (code,)).fetchone()
            if not row:
                return False, "Kod topilmadi."
            if row["is_used"]:
                return False, "Bu kod allaqachon ishlatilgan."
            conn.execute(
                "UPDATE redeem_codes SET is_used = 1, used_by_tg_user_id = ?, used_at = ? WHERE code = ?",
                (tg_user_id, utc_now(), code),
            )
            conn.commit()
            return True, "Kod aktivlashtirildi."

    def get_redeemed_code_for_user(self, code: str, tg_user_id: int) -> Optional[sqlite3.Row]:
        with closing(self.connect()) as conn:
            return conn.execute(
                """
                SELECT rc.*, e.title, e.grade_level, e.subject, e.category, e.duration_minutes, e.description,
                       e.pdf_file_id, e.answer_key, e.answer_instructions, e.is_free_demo
                FROM redeem_codes rc
                JOIN exams e ON e.id = rc.exam_id
                WHERE rc.code = ? AND rc.is_used = 1 AND rc.used_by_tg_user_id = ? AND e.is_active = 1
                """,
                (normalize_code(code), tg_user_id),
            ).fetchone()

    def save_attempt(self, tg_user_id: int, exam_id: int, code: str, user_answers_tokens: list[str], result: dict) -> None:
        with closing(self.connect()) as conn:
            conn.execute(
                """
                INSERT INTO attempts (tg_user_id, exam_id, code, user_answers, score, total_questions, percentage, correct_count, wrong_count, blank_count, five_grade, result_text, created_at)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (tg_user_id, exam_id, normalize_code(code), serialize_tokens(user_answers_tokens), result["correct_count"], result["total_questions"], result["percentage"], result["correct_count"], result["wrong_count"], result["blank_count"], result["five_grade"], result["result_text"], utc_now()),
            )
            conn.commit()

    def list_user_attempts(self, tg_user_id: int, limit: int = 12) -> list[sqlite3.Row]:
        with closing(self.connect()) as conn:
            return conn.execute(
                """
                SELECT a.*, e.title, e.subject, e.category, e.grade_level
                FROM attempts a
                JOIN exams e ON e.id = a.exam_id
                WHERE a.tg_user_id = ?
                ORDER BY a.created_at DESC
                LIMIT ?
                """,
                (tg_user_id, limit),
            ).fetchall()

    def list_attempts(self) -> list[sqlite3.Row]:
        with closing(self.connect()) as conn:
            return conn.execute(
                """
                SELECT a.*, u.full_name, u.phone, u.username, u.target_grade, e.title, e.subject, e.category, e.grade_level
                FROM attempts a
                LEFT JOIN users u ON u.tg_user_id = a.tg_user_id
                LEFT JOIN exams e ON e.id = a.exam_id
                ORDER BY a.created_at DESC
                """
            ).fetchall()

    def get_stats(self) -> dict:
        with closing(self.connect()) as conn:
            total_users = conn.execute("SELECT COUNT(*) FROM users").fetchone()[0]
            total_exams = conn.execute("SELECT COUNT(*) FROM exams WHERE is_active = 1").fetchone()[0]
            free_exams = conn.execute("SELECT COUNT(*) FROM exams WHERE is_active = 1 AND is_free_demo = 1").fetchone()[0]
            total_codes = conn.execute("SELECT COUNT(*) FROM redeem_codes").fetchone()[0]
            used_codes = conn.execute("SELECT COUNT(*) FROM redeem_codes WHERE is_used = 1").fetchone()[0]
            total_attempts = conn.execute("SELECT COUNT(*) FROM attempts").fetchone()[0]
            avg_percent = conn.execute("SELECT AVG(percentage) FROM attempts").fetchone()[0]
            return {
                "total_users": total_users,
                "total_exams": total_exams,
                "free_exams": free_exams,
                "total_codes": total_codes,
                "used_codes": used_codes,
                "unused_codes": max(total_codes - used_codes, 0),
                "total_attempts": total_attempts,
                "avg_percent": round(float(avg_percent), 1) if avg_percent is not None else 0.0,
            }

    def add_message_history(self, tg_user_id: int, chat_id: int, message_id: int, sender_type: str) -> None:
        with closing(self.connect()) as conn:
            conn.execute(
                "INSERT INTO message_history (tg_user_id, chat_id, message_id, sender_type, created_at) VALUES (?, ?, ?, ?, ?)",
                (tg_user_id, chat_id, message_id, sender_type, utc_now()),
            )
            conn.commit()

    def get_message_history(self, tg_user_id: int, chat_id: int) -> list[sqlite3.Row]:
        with closing(self.connect()) as conn:
            return conn.execute("SELECT * FROM message_history WHERE tg_user_id = ? AND chat_id = ? ORDER BY id DESC LIMIT 300", (tg_user_id, chat_id)).fetchall()

    def clear_message_history(self, tg_user_id: int, chat_id: int) -> None:
        with closing(self.connect()) as conn:
            conn.execute("DELETE FROM message_history WHERE tg_user_id = ? AND chat_id = ?", (tg_user_id, chat_id))
            conn.commit()


SETTINGS = Settings(
    bot_token=os.getenv("BOT_TOKEN", "").strip(),
    admin_ids={int(x) for x in os.getenv("ADMIN_IDS", "").split(",") if x.strip().isdigit()},
    db_path=resolve_db_path(os.getenv("DB_PATH", str(BASE_DIR / "data/data.sqlite3")).strip()),
    bot_brand_name=os.getenv("BOT_BRAND_NAME", "Maktab imtihon bot").strip(),
    help_admin_url=os.getenv("HELP_ADMIN_URL", "").strip(),
    help_admin_label=os.getenv("HELP_ADMIN_LABEL", "Admin bilan bog'lanish").strip(),
)
ensure_parent_dir(SETTINGS.db_path)
DB = Database(SETTINGS.db_path)
router = Router()


def is_admin(user_id: int) -> bool:
    return user_id in SETTINGS.admin_ids


def main_menu_options(user_id: int) -> list[str]:
    options = USER_MENU_OPTIONS.copy()
    if is_admin(user_id):
        options.insert(-1, BTN_ADMIN)
    return options


def main_menu_text(user_id: int) -> str:
    return numbered_prompt("🏠 Asosiy menu", main_menu_options(user_id))


def admin_menu_text() -> str:
    return numbered_prompt("🛠 Admin panel", ADMIN_MENU_OPTIONS)


def remember_numbered_options(state: FSMContext, options: list[str]):
    return state.update_data(numbered_options=options)


def back_keyboard(rows: list[list[str]]) -> ReplyKeyboardMarkup:
    keyboard = [[KeyboardButton(text=item) for item in row] for row in rows]
    keyboard.append([KeyboardButton(text=BTN_BACK), KeyboardButton(text=BTN_HOME)])
    return ReplyKeyboardMarkup(keyboard=keyboard, resize_keyboard=True)


def main_menu(user_id: int) -> ReplyKeyboardMarkup:
    rows = [
        [BTN_FULL_MOCK, BTN_SUBJECTS],
        [BTN_CHECK, BTN_PROFILE],
        [BTN_GUIDE, BTN_HELP],
    ]
    if is_admin(user_id):
        rows.append([BTN_ADMIN])
    rows.append([BTN_CLEAR])
    return ReplyKeyboardMarkup(keyboard=[[KeyboardButton(text=item) for item in row] for row in rows], resize_keyboard=True, input_field_placeholder="Bo'limni tanlang")


def admin_menu() -> ReplyKeyboardMarkup:
    return ReplyKeyboardMarkup(
        keyboard=[
            [KeyboardButton(text=BTN_EXAM_ADD), KeyboardButton(text=BTN_EXAM_DELETE)],
            [KeyboardButton(text=BTN_EXAM_LIST), KeyboardButton(text=BTN_CODE_CREATE)],
            [KeyboardButton(text=BTN_STATS), KeyboardButton(text=BTN_BROADCAST)],
            [KeyboardButton(text=BTN_EXPORT_USERS), KeyboardButton(text=BTN_EXPORT_RESULTS)],
            [KeyboardButton(text=BTN_ADMIN_TO_USER), KeyboardButton(text=BTN_CLEAR)],
        ],
        resize_keyboard=True,
    )


def contact_keyboard() -> ReplyKeyboardMarkup:
    return ReplyKeyboardMarkup(keyboard=[[KeyboardButton(text="📱 Kontakt yuborish", request_contact=True)]], resize_keyboard=True, one_time_keyboard=True)


def grade_keyboard() -> ReplyKeyboardMarkup:
    return ReplyKeyboardMarkup(keyboard=[[KeyboardButton(text="9-sinf"), KeyboardButton(text="11-sinf")], [KeyboardButton(text="Hozircha bilmayman")]], resize_keyboard=True, one_time_keyboard=True)


def categories_keyboard() -> ReplyKeyboardMarkup:
    return back_keyboard([[CATEGORY_EXACT, CATEGORY_NATURAL], [CATEGORY_SOCIAL]])


def subjects_keyboard() -> ReplyKeyboardMarkup:
    return back_keyboard([[SUBJ_MATH, SUBJ_ENGLISH], [SUBJ_NATIVE, SUBJ_HISTORY], [SUBJ_GEOGRAPHY, SUBJ_BIOLOGY], [SUBJ_CHEMISTRY, SUBJ_PHYSICS]])


def subject_action_keyboard(user_id: int, grade_level: str, subject: str) -> ReplyKeyboardMarkup:
    free_used = DB.has_used_free_demo(user_id, grade_level if grade_level in {"9", "11"} else "UNKNOWN", subject)
    first_row = [BTN_ENTER_CODE]
    if not free_used:
        first_row.append(BTN_FREE_TEST)
    return back_keyboard([first_row, [BTN_TESTS]])


def full_mock_action_keyboard() -> ReplyKeyboardMarkup:
    return back_keyboard([[BTN_ENTER_CODE], [BTN_FREE_TEST]])


def answer_examples_text(total_questions: int) -> str:
    example_count = min(max(total_questions, 4), 8)
    base_letters = ["A", "B", "C", "D", "A", "C", "B", "D"][:example_count]
    return (
        "Javob kiritish formatlari:\n"
        f"• {''.join(base_letters)}\n"
        f"• {' '.join(base_letters)}\n"
        f"• {', '.join(f'{i + 1}-{ans}' for i, ans in enumerate(base_letters))}\n"
        "Savollar tartibini buzmasdan yuboring."
    )


def school_exam_info_text() -> str:
    return (
        "📘 Botdagi bo'limlar\n\n"
        "🎯 Full mock — faqat kod kiritish va bir martalik bonus mock olish bo'limi.\n"
        "📚 Fanlar — alohida fan bo'yicha testlar, kod orqali test olish va mavjud testlar ro'yxati.\n"
        "✅ Test tekshirish — ishlangan testni kod va javoblar orqali tez tekshirish.\n"
        "👤 Shaxsiy ma'lumotlar — telefoningiz, ishlatgan kodlaringiz va oxirgi natijalaringiz.\n\n"
        "Full mock ichida fan tanlanmaydi: user faqat kod kiritadi yoki admin qo'shgan bir martalik bonus mockni oladi. Bonus mock har user uchun faqat 1 marta ochiladi."
    )


def help_text() -> str:
    lines = [
        "🧭 Botdan foydalanish qo'llanmasi",
        "",
        "1) /start bosing va ro'yxatdan o'ting.",
        "2) Full mock bo'limida kod kiriting yoki bir martalik bonus mockni oling.",
        "3) Fanlar bo'limida alohida fanlarni ko'rishingiz mumkin.",
        "4) Testni ishlab bo'lgach ✅ Test tekshirish bo'limiga kiring.",
        "5) Kodni yuboring, so'ng javoblarni yuboring.",
        "6) Natija foiz, baho va tahlil bilan chiqadi.",
        "7) Shaxsiy ma'lumotlar bo'limida ishlatgan kodlar va oxirgi natijalarni ko'rasiz.",
    ]
    if SETTINGS.help_admin_url:
        lines += ["", f"🆘 <a href=\"{escape_html_text(SETTINGS.help_admin_url)}\">{escape_html_text(SETTINGS.help_admin_label)}</a>"]
    else:
        lines += ["", DEFAULT_HELP_TEXT]
    return "\n".join(lines)


def format_exam_line(exam: sqlite3.Row) -> str:
    count = len(deserialize_tokens(exam["answer_key"]))
    demo_tag = " | FREE" if exam["is_free_demo"] else ""
    return f"ID {exam['id']} | {grade_label(exam['grade_level'])} | {exam['category']} | {exam['subject']} | {exam['title']} | {count} savol{demo_tag}"


def format_exam_numbered_list(exams: list[sqlite3.Row]) -> str:
    return "\n".join(f"{index}) {format_exam_line(exam)}" for index, exam in enumerate(exams, start=1))


def pick_exam_id_from_number(text: str, exam_ids: list[int]) -> int:
    raw = compact_spaces(text)
    if not raw.isdigit():
        return 0
    num = int(raw)
    if 1 <= num <= len(exam_ids):
        return int(exam_ids[num - 1])
    if DB.get_exam(num):
        return num
    return 0


def format_exam_caption(row: sqlite3.Row, code: Optional[str] = None) -> str:
    count = len(deserialize_tokens(row["answer_key"]))
    parts = [
        f"📄 <b>{escape_html_text(row['title'])}</b>",
        f"Sinf: {escape_html_text(grade_label(row['grade_level']))}",
        f"Kategoriya: {escape_html_text(row['category'] or '-')}",
        f"Fan: {escape_html_text(row['subject'] or '-')}",
        f"Savollar: {count}",
        f"Vaqt: {row['duration_minutes']} daqiqa",
        f"Izoh: {escape_html_text(row['description'] or '-')}",
    ]
    if code:
        parts.append(f"Kod: <code>{escape_html_text(code)}</code>")
    parts += ["", "Ishlab bo'lgach <b>✅ Test tekshirish</b> bo'limiga kiring."]
    return "\n".join(parts)


def parse_positive_int(text: str, default: int = DEFAULT_DURATION) -> int:
    raw = compact_spaces(text)
    if raw in {"-", "", "default", "standart"}:
        return default
    try:
        value = int(raw)
        return value if value > 0 else default
    except ValueError:
        return default


def get_subject_context(data: dict) -> tuple[str, str, str]:
    return data.get("selected_grade", "UNKNOWN"), data.get("selected_category", ""), data.get("selected_subject", "")


def normalize_category(text: str) -> str:
    text = compact_spaces(text)
    return text if text in CATEGORY_SUBJECTS else ""


def normalize_subject(text: str) -> str:
    text = compact_spaces(text)
    return text if text in SUBJECTS else ""


def profile_text(user: sqlite3.Row, attempts: list[sqlite3.Row], codes: list[sqlite3.Row]) -> str:
    lines = [
        "👤 Shaxsiy ma'lumotlaringiz",
        "",
        f"Ism: {escape_html_text(user['full_name'] or '-')}",
        f"Username: @{escape_html_text(user['username'])}" if user["username"] else "Username: -",
        f"Telefon: {escape_html_text(user['phone'] or '-')}",
        f"Ro'yxatdan o'tgan sana: {escape_html_text(user['registered_at'])}",
        "",
        f"Oxirgi natijalar soni: {len(attempts)}",
    ]
    if attempts:
        lines.append("So'nggi natijalar:")
        for row in attempts[:5]:
            lines.append(f"• {escape_html_text(row['subject'])} | {row['correct_count']}/{row['total_questions']} | {row['percentage']}% | baho {row['five_grade']}")
    else:
        lines.append("So'nggi natijalar: hali yo'q")
    lines += ["", "Ishlatgan kodlaringiz:"]
    if codes:
        for row in codes[:10]:
            lines.append(f"• <code>{escape_html_text(row['code'])}</code> | {escape_html_text(row['subject'])} | {escape_html_text(row['title'])}")
    else:
        lines.append("• Hali ishlatilgan kod yo'q")
    return "\n".join(lines)


def export_users_xlsx() -> Path:
    rows = []
    for idx, row in enumerate(DB.list_users(), start=1):
        rows.append([idx, row["tg_user_id"], row["username"] or "-", row["full_name"] or "-", row["phone"] or "-", row["attempts_count"], row["registered_at"]])
    return build_xlsx_file("users_export.xlsx", ["#", "TG ID", "Username", "F.I.Sh", "Telefon", "Urinishlar", "Sana"], rows, "Userlar export")


def export_users_pdf() -> Path:
    rows = []
    for idx, row in enumerate(DB.list_users(), start=1):
        rows.append([idx, row["tg_user_id"], row["username"] or "-", row["full_name"] or "-", row["phone"] or "-", row["attempts_count"], row["registered_at"]])
    return build_pdf_file("users_export.pdf", "Userlar export", ["#", "TG ID", "Username", "F.I.Sh", "Telefon", "Urinishlar", "Sana"], rows)


def export_results_xlsx() -> Path:
    rows = []
    for idx, row in enumerate(DB.list_attempts(), start=1):
        rows.append([idx, row["full_name"] or "-", row["phone"] or "-", grade_label(row["target_grade"]), row["category"] or "-", row["subject"] or "-", row["title"] or "-", row["code"], row["correct_count"], row["total_questions"], row["percentage"], row["five_grade"], row["created_at"]])
    return build_xlsx_file("results_export.xlsx", ["#", "F.I.Sh", "Telefon", "Sinf", "Kategoriya", "Fan", "Test", "Kod", "Togri", "Jami", "Foiz", "Baho", "Sana"], rows, "Natijalar export")


def export_results_pdf() -> Path:
    rows = []
    for idx, row in enumerate(DB.list_attempts(), start=1):
        rows.append([idx, row["full_name"] or "-", row["phone"] or "-", grade_label(row["target_grade"]), row["category"] or "-", row["subject"] or "-", row["title"] or "-", row["code"], row["correct_count"], row["total_questions"], row["percentage"], row["five_grade"], row["created_at"]])
    return build_pdf_file("results_export.pdf", "Natijalar export", ["#", "F.I.Sh", "Telefon", "Sinf", "Kategoriya", "Fan", "Test", "Kod", "Togri", "Jami", "Foiz", "Baho", "Sana"], rows)


async def track_incoming(message: Message) -> None:
    DB.add_message_history(message.from_user.id, message.chat.id, message.message_id, "user")


async def tracked_answer(message: Message, text: str, reply_markup: Optional[ReplyKeyboardMarkup] = None) -> Message:
    sent = await message.answer(text, reply_markup=reply_markup, disable_web_page_preview=True)
    DB.add_message_history(message.from_user.id, message.chat.id, sent.message_id, "bot")
    return sent


async def tracked_answer_document(message: Message, document: FSInputFile, caption: str, reply_markup: Optional[ReplyKeyboardMarkup] = None) -> Message:
    sent = await message.answer_document(document=document, caption=caption, reply_markup=reply_markup)
    DB.add_message_history(message.from_user.id, message.chat.id, sent.message_id, "bot")
    return sent


async def clear_history(message: Message, state: FSMContext) -> None:
    history = DB.get_message_history(message.from_user.id, message.chat.id)
    deleted = 0
    for row in history:
        try:
            await message.bot.delete_message(chat_id=row["chat_id"], message_id=row["message_id"])
            deleted += 1
        except TelegramBadRequest:
            continue
    DB.clear_message_history(message.from_user.id, message.chat.id)
    await state.clear()
    sent = await message.answer(f"🧹 Tozalandi. O'chirilgan xabarlar: {deleted}", reply_markup=main_menu(message.from_user.id))
    DB.add_message_history(message.from_user.id, message.chat.id, sent.message_id, "bot")


async def show_subject_hub(message: Message, state: FSMContext, subject: str, category: str) -> None:
    user = DB.get_user(message.from_user.id)
    grade = user["target_grade"] if user else "UNKNOWN"
    exams = DB.list_exams(subject=subject, grade_level=grade, category=category, include_free=True)
    normal_count = len([x for x in exams if not x["is_free_demo"]])
    free_used = DB.has_used_free_demo(message.from_user.id, grade if grade in {"9", "11"} else "UNKNOWN", subject)
    action_options = [BTN_ENTER_CODE]
    if not free_used:
        action_options.append(BTN_FREE_TEST)
    action_options.append(BTN_TESTS)
    await state.update_data(selected_subject=subject, selected_category=category, selected_grade=grade, numbered_options=action_options)
    lines = [
        f"📘 {subject}",
        f"Kategoriya: {category}",
        "",
        f"Mavjud testlar: {normal_count}",
        f"Bir martalik tekin test: {'ishlatilgan' if free_used else 'mavjud bo\'lishi mumkin'}",
        "",
        numbered_prompt("Quyidan kerakli amalni tanlang:", action_options),
    ]
    await tracked_answer(message, "\n".join(lines), reply_markup=subject_action_keyboard(message.from_user.id, grade if grade else "UNKNOWN", subject))


async def send_exam_document(message: Message, exam_row: sqlite3.Row, code_text: str, reply_markup: Optional[ReplyKeyboardMarkup] = None) -> None:
    try:
        sent = await message.answer_document(document=exam_row["pdf_file_id"], caption=format_exam_caption(exam_row, code_text), reply_markup=reply_markup)
        DB.add_message_history(message.from_user.id, message.chat.id, sent.message_id, "bot")
    except TelegramBadRequest:
        await tracked_answer(message, format_exam_caption(exam_row, code_text), reply_markup=reply_markup)


@router.message(CommandStart())
async def start_handler(message: Message, state: FSMContext) -> None:
    await track_incoming(message)
    await state.clear()
    DB.upsert_user(message.from_user.id, username=message.from_user.username)
    user = DB.get_user(message.from_user.id)
    if user and user["full_name"] and user["phone"]:
        await remember_numbered_options(state, main_menu_options(message.from_user.id))
        await tracked_answer(message, f"{SETTINGS.bot_brand_name} ga xush kelibsiz.\n\n" + main_menu_text(message.from_user.id), reply_markup=main_menu(message.from_user.id))
        return
    await state.set_state(RegisterStates.waiting_name)
    await tracked_answer(message, "Assalomu alaykum. Ism-familiyangizni yuboring:", reply_markup=ReplyKeyboardRemove())


@router.message(Command("admin"))
async def admin_handler(message: Message, state: FSMContext) -> None:
    await track_incoming(message)
    if not is_admin(message.from_user.id):
        await tracked_answer(message, "Siz admin emassiz.", reply_markup=main_menu(message.from_user.id))
        return
    await state.clear()
    await remember_numbered_options(state, ADMIN_MENU_OPTIONS)
    await tracked_answer(message, admin_menu_text(), reply_markup=admin_menu())


@router.message(RegisterStates.waiting_name, F.content_type == ContentType.TEXT)
async def registration_name_handler(message: Message, state: FSMContext) -> None:
    await track_incoming(message)
    full_name = compact_spaces(message.text or "")
    if not is_valid_full_name(full_name):
        await tracked_answer(message, "Ism-familiya noto'g'ri. Raqam ishlatmang va faqat harflar bilan qayta yuboring:")
        return
    DB.upsert_user(message.from_user.id, username=message.from_user.username, full_name=full_name)
    await state.set_state(RegisterStates.waiting_contact)
    await tracked_answer(message, "Endi telefon raqamingizni yuboring:", reply_markup=contact_keyboard())


@router.message(RegisterStates.waiting_contact, F.contact)
async def registration_contact_handler(message: Message, state: FSMContext) -> None:
    await track_incoming(message)
    if not message.contact or message.contact.user_id != message.from_user.id:
        await tracked_answer(message, "Faqat o'zingizning kontaktingizni tugma orqali yuboring:", reply_markup=contact_keyboard())
        return
    phone = message.contact.phone_number if message.contact else ""
    if not is_valid_phone(phone):
        await tracked_answer(message, "Kontakt noto'g'ri. Tugma orqali qayta yuboring:", reply_markup=contact_keyboard())
        return
    DB.upsert_user(message.from_user.id, phone=phone, target_grade="UNKNOWN")
    await state.clear()
    await remember_numbered_options(state, main_menu_options(message.from_user.id))
    await tracked_answer(message, "✅ Ro'yxatdan o'tish tugadi.\n\n" + main_menu_text(message.from_user.id), reply_markup=main_menu(message.from_user.id))


@router.message(RegisterStates.waiting_contact, F.content_type == ContentType.TEXT)
async def registration_contact_text_handler(message: Message, state: FSMContext) -> None:
    await track_incoming(message)
    await tracked_answer(
        message,
        "Telefon raqam qo'lda yozib qabul qilinmaydi. Pastdagi tugma orqali kontakt yuboring:",
        reply_markup=contact_keyboard(),
    )


@router.message(RegisterStates.waiting_contact)
async def registration_contact_fallback_handler(message: Message, state: FSMContext) -> None:
    await track_incoming(message)
    await tracked_answer(message, "Faqat pastdagi tugma orqali kontakt yuboring:", reply_markup=contact_keyboard())


@router.message(RegisterStates.waiting_grade, F.content_type == ContentType.TEXT)
async def registration_grade_handler(message: Message, state: FSMContext) -> None:
    await track_incoming(message)
    grade = GRADE_BUTTONS.get(compact_spaces(message.text or ""), "")
    if not grade:
        await tracked_answer(message, "Pastdagi tugmalardan birini tanlang:", reply_markup=grade_keyboard())
        return
    DB.upsert_user(message.from_user.id, target_grade=grade)
    await state.clear()
    await remember_numbered_options(state, main_menu_options(message.from_user.id))
    await tracked_answer(message, "✅ Ro'yxatdan o'tish tugadi.\n\n" + main_menu_text(message.from_user.id), reply_markup=main_menu(message.from_user.id))


@router.message(AdminStates.waiting_exam_title, F.content_type == ContentType.TEXT)
async def admin_exam_title(message: Message, state: FSMContext) -> None:
    await track_incoming(message)
    title = compact_spaces(message.text or "")
    if len(title) < 3:
        await tracked_answer(message, "Nom juda qisqa. Qayta yuboring:")
        return
    await state.update_data(exam_title=title)
    await state.set_state(AdminStates.waiting_exam_grade)
    await tracked_answer(message, numbered_prompt("Test qaysi sinf uchun?", [label for label, _ in GRADE_OPTIONS]), reply_markup=ReplyKeyboardRemove())


@router.message(AdminStates.waiting_exam_grade, F.content_type == ContentType.TEXT)
async def admin_exam_grade(message: Message, state: FSMContext) -> None:
    await track_incoming(message)
    grade = pick_labeled_value(message.text or "", GRADE_OPTIONS) or normalize_grade_value(message.text or "")
    if not grade:
        await tracked_answer(message, numbered_prompt("Faqat quyidagilardan birini tanlang:", [label for label, _ in GRADE_OPTIONS]))
        return
    await state.update_data(exam_grade=grade)
    await state.set_state(AdminStates.waiting_exam_category)
    await tracked_answer(message, numbered_prompt("Kategoriya tanlang:", list(CATEGORY_SUBJECTS.keys())))


@router.message(AdminStates.waiting_exam_category, F.content_type == ContentType.TEXT)
async def admin_exam_category(message: Message, state: FSMContext) -> None:
    await track_incoming(message)
    category = pick_numbered_value(message.text or "", list(CATEGORY_SUBJECTS.keys())) or normalize_category(message.text or "")
    if not category:
        await tracked_answer(message, numbered_prompt("Kategoriya noto'g'ri. Tayyor variantlardan birini tanlang:", list(CATEGORY_SUBJECTS.keys())))
        return
    await state.update_data(exam_category=category)
    await state.set_state(AdminStates.waiting_exam_subject)
    await tracked_answer(message, numbered_prompt("Fan nomini tanlang:", CATEGORY_SUBJECTS[category]))


@router.message(AdminStates.waiting_exam_subject, F.content_type == ContentType.TEXT)
async def admin_exam_subject(message: Message, state: FSMContext) -> None:
    await track_incoming(message)
    data = await state.get_data()
    category = data.get("exam_category", "")
    subject = pick_numbered_value(message.text or "", CATEGORY_SUBJECTS.get(category, [])) or normalize_subject(message.text or "")
    if not subject or subject not in CATEGORY_SUBJECTS.get(category, []):
        await tracked_answer(message, numbered_prompt("Fan noto'g'ri yoki kategoriya bilan mos emas. Qayta tanlang:", CATEGORY_SUBJECTS.get(category, [])))
        return
    await state.update_data(exam_subject=subject)
    await state.set_state(AdminStates.waiting_exam_duration)
    await tracked_answer(message, f"Davomiylikni daqiqada yuboring. Masalan 180. Bo'sh qoldirsangiz {DEFAULT_DURATION} bo'ladi.")


@router.message(AdminStates.waiting_exam_duration, F.content_type == ContentType.TEXT)
async def admin_exam_duration(message: Message, state: FSMContext) -> None:
    await track_incoming(message)
    await state.update_data(exam_duration=parse_positive_int(message.text or "", DEFAULT_DURATION))
    await state.set_state(AdminStates.waiting_exam_description)
    await tracked_answer(message, "Qisqa izoh yuboring. Kerak bo'lmasa '-' yuboring.")


@router.message(AdminStates.waiting_exam_description, F.content_type == ContentType.TEXT)
async def admin_exam_description(message: Message, state: FSMContext) -> None:
    await track_incoming(message)
    desc = compact_spaces(message.text or "")
    if desc == "-":
        desc = ""
    await state.update_data(exam_description=desc)
    await state.set_state(AdminStates.waiting_exam_free_flag)
    await tracked_answer(message, numbered_prompt("Bu test bir martalik tekin testmi?", [label for label, _ in YES_NO_OPTIONS]))


@router.message(AdminStates.waiting_exam_free_flag, F.content_type == ContentType.TEXT)
async def admin_exam_free_flag(message: Message, state: FSMContext) -> None:
    await track_incoming(message)
    raw = compact_spaces(message.text or "").lower()
    choice = pick_labeled_value(message.text or "", YES_NO_OPTIONS)
    if choice == "yes" or raw in {"ha", "yes"}:
        is_free = True
    elif choice == "no" or raw in {"yo'q", "yoq", "no", "0"}:
        is_free = False
    else:
        await tracked_answer(message, numbered_prompt("Faqat Ha yoki Yo'q tanlang:", [label for label, _ in YES_NO_OPTIONS]))
        return
    await state.update_data(exam_is_free=is_free)
    await state.set_state(AdminStates.waiting_exam_pdf)
    await tracked_answer(message, "Endi test PDF faylini yuboring.")


@router.message(AdminStates.waiting_exam_pdf, F.document)
async def admin_exam_pdf(message: Message, state: FSMContext) -> None:
    await track_incoming(message)
    if not is_pdf_document(message):
        await tracked_answer(message, "Faqat PDF fayl yuboring.")
        return
    await state.update_data(exam_pdf_file_id=message.document.file_id)
    await state.set_state(AdminStates.waiting_exam_answer_key)
    await tracked_answer(message, "Javoblar kalitini yuboring. Masalan: ABCDABCD yoki A B C D yoki 1-A, 2-B")


@router.message(AdminStates.waiting_exam_pdf)
async def admin_exam_pdf_fallback(message: Message, state: FSMContext) -> None:
    await track_incoming(message)
    await tracked_answer(message, "PDF fayl yuboring.")


@router.message(AdminStates.waiting_exam_answer_key, F.content_type == ContentType.TEXT)
async def admin_exam_answer_key(message: Message, state: FSMContext) -> None:
    await track_incoming(message)
    tokens = parse_answer_tokens(message.text or "")
    if len(tokens) < 2:
        await tracked_answer(message, "Kamida 2 ta javob bo'lishi kerak. Qayta yuboring.")
        return
    await state.update_data(exam_answer_key=tokens)
    await state.set_state(AdminStates.waiting_exam_instructions)
    await tracked_answer(message, "Javob yuborish yo'riqnomasini yozing yoki '-' yuboring.")


@router.message(AdminStates.waiting_exam_instructions, F.content_type == ContentType.TEXT)
async def admin_exam_instructions(message: Message, state: FSMContext) -> None:
    await track_incoming(message)
    data = await state.get_data()
    instructions = compact_spaces(message.text or "")
    if instructions == "-":
        instructions = ""
    exam_id = DB.create_exam(
        title=data["exam_title"],
        grade_level=data["exam_grade"],
        category=data["exam_category"],
        subject=data["exam_subject"],
        duration_minutes=int(data["exam_duration"]),
        description=data.get("exam_description", ""),
        pdf_file_id=data["exam_pdf_file_id"],
        answer_key_tokens=data["exam_answer_key"],
        answer_instructions=instructions,
        is_free_demo=bool(data.get("exam_is_free", False)),
    )
    await state.clear()
    await remember_numbered_options(state, ADMIN_MENU_OPTIONS)
    await tracked_answer(message, f"✅ Test qo'shildi. ID: {exam_id}\n\n" + admin_menu_text(), reply_markup=admin_menu())


@router.message(AdminStates.waiting_delete_exam_id, F.content_type == ContentType.TEXT)
async def admin_delete_exam(message: Message, state: FSMContext) -> None:
    await track_incoming(message)
    data = await state.get_data()
    exam_ids = [int(x) for x in data.get("delete_exam_ids", [])]
    exam_id = pick_exam_id_from_number(message.text or "", exam_ids)
    if not exam_id:
        await tracked_answer(message, "Faqat ro'yxatdagi raqamni yuboring. Masalan: 2")
        return
    ok = DB.deactivate_exam(exam_id)
    await state.clear()
    await remember_numbered_options(state, ADMIN_MENU_OPTIONS)
    await tracked_answer(message, ("✅ O'chirildi." if ok else "ID topilmadi.") + "\n\n" + admin_menu_text(), reply_markup=admin_menu())


@router.message(AdminStates.waiting_code_exam_id, F.content_type == ContentType.TEXT)
async def admin_code_exam_id(message: Message, state: FSMContext) -> None:
    await track_incoming(message)
    data = await state.get_data()
    exam_ids = [int(x) for x in data.get("code_exam_ids", [])]
    exam_id = pick_exam_id_from_number(message.text or "", exam_ids)
    if not exam_id:
        await tracked_answer(message, "To'g'ri ro'yxat raqamini yuboring. Masalan: 2")
        return
    await state.update_data(code_exam_id=exam_id)
    await state.set_state(AdminStates.waiting_code_mode)
    await tracked_answer(message, numbered_prompt("Kod yaratish turini tanlang:", [label for label, _ in CODE_MODE_OPTIONS]))


@router.message(AdminStates.waiting_code_mode, F.content_type == ContentType.TEXT)
async def admin_code_mode(message: Message, state: FSMContext) -> None:
    await track_incoming(message)
    raw = pick_labeled_value(message.text or "", CODE_MODE_OPTIONS)
    if raw == "1":
        await state.set_state(AdminStates.waiting_single_code)
        await tracked_answer(message, "Kodning o'zini yuboring.")
        return
    if raw == "2":
        await state.set_state(AdminStates.waiting_bulk_codes)
        await tracked_answer(message, "Kodlarni bo'shliq yoki qator bilan yuboring.")
        return
    await tracked_answer(message, numbered_prompt("Faqat quyidagilardan birini tanlang:", [label for label, _ in CODE_MODE_OPTIONS]))


@router.message(AdminStates.waiting_single_code, F.content_type == ContentType.TEXT)
async def admin_single_code(message: Message, state: FSMContext) -> None:
    await track_incoming(message)
    data = await state.get_data()
    code = normalize_code(message.text or "")
    if not code:
        await tracked_answer(message, "Kod bo'sh bo'lmasin.")
        return
    ok = DB.create_code(code, int(data["code_exam_id"]), message.from_user.id)
    await state.clear()
    await remember_numbered_options(state, ADMIN_MENU_OPTIONS)
    await tracked_answer(message, f"{'✅ Yaratildi' if ok else 'Bu kod mavjud'}: {code}\n\n" + admin_menu_text(), reply_markup=admin_menu())


@router.message(AdminStates.waiting_bulk_codes, F.content_type == ContentType.TEXT)
async def admin_bulk_codes(message: Message, state: FSMContext) -> None:
    await track_incoming(message)
    data = await state.get_data()
    raw_codes = re.split(r"[\s,;|]+", message.text or "")
    created, skipped = DB.bulk_create_codes(raw_codes, int(data["code_exam_id"]), message.from_user.id)
    await state.clear()
    await remember_numbered_options(state, ADMIN_MENU_OPTIONS)
    await tracked_answer(message, f"✅ Yaratildi: {created}\n⏭ O'tkazildi: {skipped}\n\n" + admin_menu_text(), reply_markup=admin_menu())


@router.message(AdminStates.waiting_export_users_format, F.content_type == ContentType.TEXT)
async def admin_export_users(message: Message, state: FSMContext) -> None:
    await track_incoming(message)
    raw = compact_spaces(message.text or "")
    raw = pick_labeled_value(message.text or "", EXPORT_FORMAT_OPTIONS) or raw
    if raw not in {"1", "2"}:
        await tracked_answer(message, numbered_prompt("Formatni tanlang:", [label for label, _ in EXPORT_FORMAT_OPTIONS]))
        return
    path = export_users_xlsx() if raw == "1" else export_users_pdf()
    await state.clear()
    await tracked_answer_document(message, FSInputFile(path), "📤 Userlar export", reply_markup=admin_menu())


@router.message(AdminStates.waiting_export_results_format, F.content_type == ContentType.TEXT)
async def admin_export_results(message: Message, state: FSMContext) -> None:
    await track_incoming(message)
    raw = compact_spaces(message.text or "")
    raw = pick_labeled_value(message.text or "", EXPORT_FORMAT_OPTIONS) or raw
    if raw not in {"1", "2"}:
        await tracked_answer(message, numbered_prompt("Formatni tanlang:", [label for label, _ in EXPORT_FORMAT_OPTIONS]))
        return
    path = export_results_xlsx() if raw == "1" else export_results_pdf()
    await state.clear()
    await tracked_answer_document(message, FSInputFile(path), "📥 Natijalar export", reply_markup=admin_menu())


@router.message(AdminStates.waiting_broadcast_text, F.content_type == ContentType.TEXT)
async def admin_broadcast(message: Message, state: FSMContext) -> None:
    await track_incoming(message)
    users = DB.list_users()
    text = compact_spaces(message.text or "")
    sent_count = 0
    for user in users:
        try:
            await message.bot.send_message(user["tg_user_id"], f"📣 Admin xabari\n\n{text}")
            sent_count += 1
        except Exception:
            continue
    await state.clear()
    await remember_numbered_options(state, ADMIN_MENU_OPTIONS)
    await tracked_answer(message, f"✅ Xabar yuborildi: {sent_count} ta user\n\n" + admin_menu_text(), reply_markup=admin_menu())


@router.message(UserStates.waiting_redeem_code, F.content_type == ContentType.TEXT)
async def user_redeem_code_handler(message: Message, state: FSMContext) -> None:
    await track_incoming(message)
    raw_text = compact_spaces(message.text or "")
    if raw_text in INTERRUPTIBLE_USER_TEXTS:
        await state.clear()
        await text_router(message, state)
        return
    code = normalize_code(raw_text)
    if not code:
        await tracked_answer(message, "Kod bo'sh. Qayta yuboring.")
        return
    row = DB.get_code_with_exam(code)
    if not row:
        await tracked_answer(message, "Kod topilmadi.", reply_markup=main_menu(message.from_user.id))
        return
    data = await state.get_data()
    selected_subject = data.get("selected_subject")
    selected_category = data.get("selected_category")
    if selected_subject and row["subject"] != selected_subject:
        await tracked_answer(message, f"Bu kod {selected_subject} uchun emas. Kod fan bilan mos bo'lishi kerak.")
        return
    if selected_category and row["category"] != selected_category:
        await tracked_answer(message, f"Bu kod {selected_category} bo'limi uchun emas.")
        return
    ok, msg = DB.redeem_code(code, message.from_user.id)
    if not ok:
        await tracked_answer(message, msg, reply_markup=main_menu(message.from_user.id))
        return
    redeemed = DB.get_redeemed_code_for_user(code, message.from_user.id)
    await state.clear()
    if redeemed:
        await send_exam_document(message, redeemed, code, reply_markup=main_menu(message.from_user.id))
    else:
        await tracked_answer(message, msg, reply_markup=main_menu(message.from_user.id))


@router.message(UserStates.waiting_check_code, F.content_type == ContentType.TEXT)
async def user_check_code_handler(message: Message, state: FSMContext) -> None:
    await track_incoming(message)
    raw_text = compact_spaces(message.text or "")
    if raw_text in INTERRUPTIBLE_USER_TEXTS:
        await state.clear()
        await text_router(message, state)
        return
    code = normalize_code(raw_text)
    row = DB.get_redeemed_code_for_user(code, message.from_user.id)
    if not row:
        await tracked_answer(message, "Bu kod siz tomonidan ishlatilmagan yoki faol emas.", reply_markup=main_menu(message.from_user.id))
        return
    answer_key = deserialize_tokens(row["answer_key"])
    instructions = row["answer_instructions"] or answer_examples_text(len(answer_key))
    await state.update_data(check_code=code, check_exam_id=row["exam_id"], answer_key=answer_key)
    await state.set_state(UserStates.waiting_answers)
    await tracked_answer(
        message,
        (
            f"📝 {row['title']}\n"
            f"Fan: {row['subject']}\n"
            f"Savollar soni: {len(answer_key)}\n"
            f"Kod: <code>{code}</code>\n\n"
            f"{instructions}\n\n{answer_examples_text(len(answer_key))}"
        ),
        reply_markup=main_menu(message.from_user.id),
    )


@router.message(UserStates.waiting_answers, F.content_type == ContentType.TEXT)
async def user_answers_handler(message: Message, state: FSMContext) -> None:
    await track_incoming(message)
    raw_text = compact_spaces(message.text or "")
    if raw_text in INTERRUPTIBLE_USER_TEXTS:
        await state.clear()
        await text_router(message, state)
        return
    data = await state.get_data()
    answer_key = data.get("answer_key", [])
    exam_id = int(data.get("check_exam_id", 0))
    code = data.get("check_code", "")
    user_tokens = parse_answer_tokens(message.text or "")
    if len(user_tokens) < 1:
        await tracked_answer(message, "Javoblar aniqlanmadi. Qayta yuboring.")
        return
    result = evaluate_answers(answer_key, user_tokens)
    DB.save_attempt(message.from_user.id, exam_id, code, user_tokens, result)
    await state.clear()
    await tracked_answer(message, result["result_text"], reply_markup=main_menu(message.from_user.id))


@router.message(F.content_type == ContentType.TEXT)
async def text_router(message: Message, state: FSMContext) -> None:
    await track_incoming(message)
    text = compact_spaces(message.text or "")
    user = DB.get_user(message.from_user.id)

    if text == BTN_CLEAR:
        await clear_history(message, state)
        return

    if not user or not user["full_name"] or not user["phone"]:
        await state.set_state(RegisterStates.waiting_name)
        await tracked_answer(message, "Avval ro'yxatdan o'ting. Ism-familiyangizni yuboring:", reply_markup=ReplyKeyboardRemove())
        return

    data = await state.get_data()
    if text.isdigit() and data.get("numbered_options"):
        options = data.get("numbered_options") or []
        index = int(text) - 1
        if 0 <= index < len(options):
            text = options[index]

    if text == BTN_HOME:
        await state.clear()
        await remember_numbered_options(state, main_menu_options(message.from_user.id))
        await tracked_answer(message, main_menu_text(message.from_user.id), reply_markup=main_menu(message.from_user.id))
        return

    if text == BTN_BACK:
        data = await state.get_data()
        if data.get("full_mock_mode"):
            await state.clear()
            await remember_numbered_options(state, main_menu_options(message.from_user.id))
            await tracked_answer(message, "⬅️ Orqaga qaytildi.\n\n" + main_menu_text(message.from_user.id), reply_markup=main_menu(message.from_user.id))
            return
        if data.get("selected_subject"):
            await state.update_data(selected_subject="")
            if data.get("from_subjects"):
                await remember_numbered_options(state, SUBJECTS)
                await tracked_answer(message, numbered_prompt("📚 Fanlar bo'limi. Fanni tanlang:", SUBJECTS), reply_markup=subjects_keyboard())
            else:
                await remember_numbered_options(state, list(CATEGORY_SUBJECTS.keys()))
                await tracked_answer(message, numbered_prompt("🎯 Full mock bo'limi. Kategoriyani tanlang:", list(CATEGORY_SUBJECTS.keys())), reply_markup=categories_keyboard())
            return
        if data.get("selected_category"):
            await state.update_data(selected_category="")
            await remember_numbered_options(state, list(CATEGORY_SUBJECTS.keys()))
            await tracked_answer(message, numbered_prompt("🎯 Full mock bo'limi. Kategoriyani tanlang:", list(CATEGORY_SUBJECTS.keys())), reply_markup=categories_keyboard())
            return
        await remember_numbered_options(state, main_menu_options(message.from_user.id))
        await tracked_answer(message, "⬅️ Orqaga qaytildi.\n\n" + main_menu_text(message.from_user.id), reply_markup=main_menu(message.from_user.id))
        return

    if text == BTN_GUIDE:
        await tracked_answer(message, help_text(), reply_markup=main_menu(message.from_user.id))
        return

    if text == BTN_HELP:
        await tracked_answer(message, school_exam_info_text() + "\n\n" + help_text(), reply_markup=main_menu(message.from_user.id))
        return

    if text == BTN_PROFILE:
        attempts = DB.list_user_attempts(message.from_user.id)
        codes = DB.list_user_codes(message.from_user.id)
        await tracked_answer(message, profile_text(user, attempts, codes), reply_markup=main_menu(message.from_user.id))
        return

    if text == BTN_CHECK:
        await state.set_state(UserStates.waiting_check_code)
        await tracked_answer(message, "Avval kodni yuboring. Bot keyin javob kiritish formatlarini ko'rsatadi.", reply_markup=main_menu(message.from_user.id))
        return

    if text == BTN_FULL_MOCK:
        options = [BTN_ENTER_CODE, BTN_FREE_TEST]
        await state.update_data(full_mock_mode=True, from_subjects=False, selected_subject="", selected_category="", selected_grade=user["target_grade"], numbered_options=options)
        await tracked_answer(
            message,
            numbered_prompt("🎯 Full mock bo'limi. Quyidagilardan birini tanlang:", options),
            reply_markup=full_mock_action_keyboard(),
        )
        return

    if text == BTN_SUBJECTS:
        await state.update_data(from_subjects=True, selected_subject="", selected_category="", selected_grade=user["target_grade"], numbered_options=SUBJECTS)
        await tracked_answer(message, numbered_prompt("📚 Fanlar bo'limi. Fanni tanlang:", SUBJECTS), reply_markup=subjects_keyboard())
        return

    if text in CATEGORY_SUBJECTS:
        subs = CATEGORY_SUBJECTS[text]
        await state.update_data(from_subjects=False, selected_category=text, selected_subject="", selected_grade=user["target_grade"], numbered_options=subs)
        rows = []
        for i in range(0, len(subs), 2):
            rows.append(subs[i:i+2])
        await tracked_answer(message, numbered_prompt(f"{text} bo'limi. Fanni tanlang:", subs), reply_markup=back_keyboard(rows))
        return

    if text in SUBJECTS:
        data = await state.get_data()
        category = data.get("selected_category") or SUBJECT_TO_CATEGORY.get(text, "")
        await show_subject_hub(message, state, text, category)
        return

    if text == BTN_TESTS:
        data = await state.get_data()
        grade, category, subject = get_subject_context(data)
        if not subject:
            await tracked_answer(message, "Avval fan tanlang.", reply_markup=main_menu(message.from_user.id))
            return
        exams = DB.list_exams(subject=subject, grade_level=grade, category=category, include_free=False)
        if not exams:
            await tracked_answer(message, f"{subject} bo'yicha hozircha kodli testlar yo'q.", reply_markup=subject_action_keyboard(message.from_user.id, grade, subject))
            return
        lines = [f"📄 {subject} bo'yicha mavjud testlar:"]
        for exam in exams[:20]:
            lines.append(f"• {escape_html_text(exam['title'])} | {grade_label(exam['grade_level'])} | {exam['duration_minutes']} min")
        lines.append("\nTestni olish uchun shu fan ichidagi 🔑 Kod kiritish tugmasidan foydalaning.")
        await tracked_answer(message, "\n".join(lines), reply_markup=subject_action_keyboard(message.from_user.id, grade, subject))
        return

    if text == BTN_FREE_TEST:
        data = await state.get_data()
        if data.get("full_mock_mode"):
            grade = user["target_grade"] if user else "UNKNOWN"
            if DB.has_used_full_mock_bonus(message.from_user.id):
                await tracked_answer(message, "Siz bir martalik bonus mockni ishlatib bo'lgansiz.", reply_markup=full_mock_action_keyboard())
                return
            exam = DB.get_latest_full_mock_bonus_exam(grade)
            if not exam:
                await tracked_answer(message, "Bonus mock hozircha admin tomonidan qo'shilmagan.", reply_markup=full_mock_action_keyboard())
                return
            gift_code = normalize_code(f"BONUS-{message.from_user.id}-{exam['id']}")
            DB.create_code(gift_code, int(exam["id"]), 0)
            DB.redeem_code(gift_code, message.from_user.id)
            DB.mark_full_mock_bonus_used(message.from_user.id, int(exam["id"]))
            await send_exam_document(message, exam, gift_code, reply_markup=full_mock_action_keyboard())
            return

        grade, category, subject = get_subject_context(data)
        if not subject:
            await tracked_answer(message, "Avval fan tanlang.", reply_markup=main_menu(message.from_user.id))
            return
        grade_for_demo = grade if grade in {"9", "11"} else "UNKNOWN"
        if DB.has_used_free_demo(message.from_user.id, grade_for_demo, subject):
            await tracked_answer(message, "Siz bu fan bo'yicha bir martalik tekin testni ishlatib bo'lgansiz.", reply_markup=subject_action_keyboard(message.from_user.id, grade_for_demo, subject))
            return
        exam = DB.get_latest_free_demo_exam(grade_for_demo if grade_for_demo != "UNKNOWN" else "ALL", category, subject)
        if not exam and grade_for_demo == "UNKNOWN":
            exam = DB.get_latest_free_demo_exam("ALL", category, subject)
        if not exam and grade_for_demo in {"9", "11"}:
            with closing(DB.connect()) as conn:
                exam = conn.execute(
                    "SELECT * FROM exams WHERE is_active=1 AND is_free_demo=1 AND category=? AND subject=? AND grade_level='ALL' ORDER BY id DESC LIMIT 1",
                    (category, subject),
                ).fetchone()
        if not exam:
            await tracked_answer(message, "Bu fan bo'yicha tekin test hozircha yuklanmagan.", reply_markup=subject_action_keyboard(message.from_user.id, grade_for_demo, subject))
            return
        gift_code = normalize_code(f"FREE-{message.from_user.id}-{exam['id']}")
        DB.create_code(gift_code, int(exam["id"]), 0)
        DB.redeem_code(gift_code, message.from_user.id)
        DB.mark_free_demo_used(message.from_user.id, exam["id"], grade_for_demo, subject)
        await send_exam_document(message, exam, gift_code, reply_markup=subject_action_keyboard(message.from_user.id, grade_for_demo, subject))
        return

    if text == BTN_ENTER_CODE:
        data = await state.get_data()
        grade, _, subject = get_subject_context(data)
        await state.update_data(selected_grade=grade, selected_subject=subject)
        await state.set_state(UserStates.waiting_redeem_code)
        await tracked_answer(message, "Bir martalik kodni yuboring:", reply_markup=main_menu(message.from_user.id))
        return

    if text == BTN_ADMIN:
        if not is_admin(message.from_user.id):
            await tracked_answer(message, "Siz admin emassiz.", reply_markup=main_menu(message.from_user.id))
            return
        await state.clear()
        await remember_numbered_options(state, ADMIN_MENU_OPTIONS)
        await tracked_answer(message, admin_menu_text(), reply_markup=admin_menu())
        return

    if not is_admin(message.from_user.id):
        await tracked_answer(message, "Tugmalardan foydalaning yoki /start bosing.", reply_markup=main_menu(message.from_user.id))
        return

    if text == BTN_ADMIN_TO_USER:
        await state.clear()
        await remember_numbered_options(state, main_menu_options(message.from_user.id))
        await tracked_answer(message, "👥 User menu ochildi.\n\n" + main_menu_text(message.from_user.id), reply_markup=main_menu(message.from_user.id))
        return

    if text == BTN_EXAM_LIST:
        exams = DB.list_exams()
        if not exams:
            await tracked_answer(message, "Hali test yo'q.", reply_markup=admin_menu())
            return
        await tracked_answer(message, "📋 Faol testlar:\n" + format_exam_numbered_list(exams[:100]), reply_markup=admin_menu())
        return

    if text == BTN_EXAM_ADD:
        await state.set_state(AdminStates.waiting_exam_title)
        await tracked_answer(message, "Yangi test nomini yuboring:", reply_markup=ReplyKeyboardRemove())
        return

    if text == BTN_EXAM_DELETE:
        exams = DB.list_exams()
        if not exams:
            await tracked_answer(message, "Faol test yo'q.", reply_markup=admin_menu())
            return
        shown_exams = exams[:100]
        await state.update_data(delete_exam_ids=[int(exam["id"]) for exam in shown_exams])
        await state.set_state(AdminStates.waiting_delete_exam_id)
        await tracked_answer(message, "O'chirish uchun ro'yxat raqamini yuboring:\n" + format_exam_numbered_list(shown_exams), reply_markup=admin_menu())
        return

    if text == BTN_CODE_CREATE:
        exams = DB.list_exams()
        if not exams:
            await tracked_answer(message, "Avval test qo'shing.", reply_markup=admin_menu())
            return
        shown_exams = exams[:100]
        await state.update_data(code_exam_ids=[int(exam["id"]) for exam in shown_exams])
        await state.set_state(AdminStates.waiting_code_exam_id)
        await tracked_answer(message, "Qaysi test uchun kod yaratiladi? Ro'yxat raqamini yuboring:\n" + format_exam_numbered_list(shown_exams), reply_markup=admin_menu())
        return

    if text == BTN_STATS:
        stats = DB.get_stats()
        await tracked_answer(message, (
            "📈 Umumiy statistika\n\n"
            f"Userlar: {stats['total_users']}\n"
            f"Faol testlar: {stats['total_exams']}\n"
            f"Free testlar: {stats['free_exams']}\n"
            f"Jami kodlar: {stats['total_codes']}\n"
            f"Ishlatilgan kodlar: {stats['used_codes']}\n"
            f"Ishlatilmagan kodlar: {stats['unused_codes']}\n"
            f"Tekshiruvlar: {stats['total_attempts']}\n"
            f"O'rtacha foiz: {stats['avg_percent']}%"
        ), reply_markup=admin_menu())
        return

    if text == BTN_EXPORT_USERS:
        await state.set_state(AdminStates.waiting_export_users_format)
        await tracked_answer(message, numbered_prompt("Userlar export formatini tanlang:", [label for label, _ in EXPORT_FORMAT_OPTIONS]), reply_markup=admin_menu())
        return

    if text == BTN_EXPORT_RESULTS:
        await state.set_state(AdminStates.waiting_export_results_format)
        await tracked_answer(message, numbered_prompt("Natijalar export formatini tanlang:", [label for label, _ in EXPORT_FORMAT_OPTIONS]), reply_markup=admin_menu())
        return

    if text == BTN_BROADCAST:
        await state.set_state(AdminStates.waiting_broadcast_text)
        await tracked_answer(message, "Barcha userlarga yuboriladigan xabarni kiriting:", reply_markup=ReplyKeyboardRemove())
        return

    await remember_numbered_options(state, ADMIN_MENU_OPTIONS)
    await tracked_answer(message, "Admin tugmalaridan foydalaning.\n\n" + admin_menu_text(), reply_markup=admin_menu())


async def main() -> None:
    if not SETTINGS.bot_token:
        raise RuntimeError("BOT_TOKEN .env faylda ko'rsatilishi shart")
    logger.info("DB fayl manzili: %s", SETTINGS.db_path)
    bot = Bot(token=SETTINGS.bot_token, default=DefaultBotProperties(parse_mode=ParseMode.HTML))
    dp = Dispatcher()
    dp.include_router(router)
    logger.info("Bot ishga tushdi")
    await dp.start_polling(bot, drop_pending_updates=True)


if __name__ == "__main__":
    try:
        asyncio.run(main())
    except (KeyboardInterrupt, SystemExit):
        logger.info("Bot to'xtatildi")
